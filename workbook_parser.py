from __future__ import annotations

from io import BytesIO
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from models import CellRecord, SheetSummary, WorkbookModel


def _classify_workbook(sheet_names: list[str], labels: list[str]) -> str:
    text = " ".join(sheet_names + labels).lower()
    signals = {
        "lbo": ["lbo", "sources and uses", "debt schedule"],
        "dcf": ["dcf", "terminal value", "wacc", "discount rate"],
        "3-statement": ["income statement", "balance sheet", "cash flow"],
        "saas": ["arr", "mrr", "cohort", "churn"],
        "project-finance": ["dscr", "construction", "operations period"],
    }
    for model_type, keywords in signals.items():
        if any(keyword in text for keyword in keywords):
            return model_type
    return "general"


def parse_workbook(upload: BinaryIO, filename: str) -> WorkbookModel:
    content = upload.read()
    workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)

    cells: dict[str, CellRecord] = {}
    formulas: dict[str, CellRecord] = {}
    labels: dict[str, str] = {}
    sheets: list[SheetSummary] = []

    for worksheet in workbook.worksheets:
        summary = SheetSummary(
            name=worksheet.title,
            max_row=0,
            max_col=0,
        )
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col_letter, row_idx = coordinate_from_string(cell.coordinate)
                summary.max_row = max(summary.max_row, row_idx)
                summary.max_col = max(summary.max_col, column_index_from_string(col_letter))
                record = CellRecord(
                    sheet=worksheet.title,
                    address=cell.coordinate,
                    value=cell.value,
                    formula=cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
                    data_type=cell.data_type,
                )
                cells[record.key] = record
                summary.value_count += 1

                if record.formula:
                    formulas[record.key] = record
                    summary.formula_count += 1
                elif isinstance(record.value, str):
                    labels[record.key] = record.value.strip()
                    summary.text_count += 1
                elif isinstance(record.value, (int, float)):
                    summary.numeric_count += 1
        sheets.append(summary)

    named_ranges = _extract_named_ranges(workbook)

    workbook_type_hint = _classify_workbook([sheet.name for sheet in sheets], list(labels.values())[:500])
    return WorkbookModel(
        filename=filename,
        sheets=sheets,
        cells=cells,
        formulas=formulas,
        labels=labels,
        named_ranges=named_ranges,
        workbook_type_hint=workbook_type_hint,
    )


def _extract_named_ranges(workbook) -> dict[str, str]:
    named_ranges: dict[str, str] = {}
    defined_names = getattr(workbook, "defined_names", None)
    if not defined_names:
        return named_ranges

    entries = None
    if hasattr(defined_names, "items"):
        try:
            entries = list(defined_names.items())
        except Exception:
            entries = None

    if entries:
        for name, defined_name in entries:
            attr_text = getattr(defined_name, "attr_text", None)
            if attr_text:
                named_ranges[name] = attr_text
        return named_ranges

    legacy_names = getattr(defined_names, "definedName", None)
    if legacy_names:
        for defined_name in legacy_names:
            attr_text = getattr(defined_name, "attr_text", None)
            name = getattr(defined_name, "name", None)
            if name and attr_text:
                named_ranges[name] = attr_text

    return named_ranges
